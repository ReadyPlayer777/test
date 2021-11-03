# 这是一个示例 Python 脚本。
'''
单元格格式说明
C2:科室名称 J2:运行日期 C3:本周运行和管理 C36:定期任务和执行 C41:交办任务 C45:下周计划 C49:点评及任务布置 C53:文件传达及学习
'''
import xlrd
import xlwt
import openpyxl
import os
from dateutil.parser import parse


def print_hi(name):
    # 在下面的代码行中使用断点来调试脚本。
    print(f'Hi, {name}')  # 按 Ctrl+F8 切换断点。


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    print_hi('PyCharm')
    # 读取当前目录下的excel文件（工作记录）
    excel_path = os.path.join(os.getcwd(), '2021年度值班工作记录表.xls')
    print("文件路径：", excel_path)
    excel_file = xlrd.open_workbook(excel_path)
    # 查看所有表格名字
    names = excel_file.sheet_names()
    print("所有表格:", names)
    # 读取第一张表格
    table = excel_file.sheets()[2]  # 这里拿2021年3月做周报例子

    # 打开周报模板
    excel_path_ZhouBao = os.path.join(os.getcwd(), 'pattern.xlsx')
    wb = openpyxl.load_workbook(excel_path_ZhouBao)
    sheet = wb['周报']
    sheet.title = '周报'
    sheet['C2'] = '终端运行室'
    '''sheet['C2'] = '终端运行室\ndd'
    wb.save('周报初稿.xlsx')'''
    # 工作记录中的行数
    nrows = table.nrows
    # 工作记录中的列数
    ncols = table.ncols
    # 周报统计的轮数,从第六行开始,第一轮
    ncount = 6
    # 行内容
    cell_content = ''
    start_date = parse(str(int(table.cell(6, 0).value))).date()  # 周报开始时间
    print('周报开始时间', start_date)
    # 遍历工作记录表，开始统计绘制周报
    for i in range(ncount, nrows):
        cell_date = parse(str(int(table.cell(i, 0).value))).date()  # 当前位置的日期
        cell_time = str(cell_date.month) + '月' + str(cell_date.day) + '日' + '\n'  # 将当前位置的日期转化为xx月xx日格式
        cell_content = cell_content + cell_time + str(table.cell(i, 4).value)
        # print(cell_content)
        if '星期四' in str(table.cell(i, 1).value):  # 周四完成一周的周报统计
            ncount = i + 1
            end_date = parse(str(int(table.cell(i, 0).value))).date()  # 周报结束时间
            sheet['C3'] = cell_content
            sheet['J2'] = str(start_date.year) + '年' + str(start_date.month) + '月' + str(start_date.day) + '日至' + str(
                end_date.year) + '年' + str(end_date.month) + '月' + str(end_date.day) + '日'
            save_name = '周报初稿' + str(start_date) + '到' + str(end_date) + '.xlsx'
            wb.save(save_name)
            break
            # print('周报结束时间', end_date)
        if ncount == nrows:
            break
    print(cell_content)
    '''for i in range(nrows):
        row_content = []
        for j in range(ncols):
            cell_value = table.cell(i, j).value
            if cell_value is not None and i != 0:
                # print(cell_value)
                row_content.append(cell_value)
        sheet.write(i, 0, row_content)
    new_file.save(excel_path_new)
    '''
    # print(table.row_values(2))
