# 这是一个示例 Python 脚本。
'''
单元格格式说明
C2:科室名称 J2:运行日期 C3:本周运行和管理 C46:定期任务和执行 C51:交办任务 C55:下周计划 C59:点评及任务布置 C63:文件传达及学习
'''
import xlrd
import xlwt
import openpyxl
import os
from dateutil.parser import parse


# 值班记录类
class Record():
    def __init__(self, date, week, content):
        self.date = date
        self.week = week
        self.content = content

    def printinfo(self):
        print("日期:", self.date)
        print("星期:", self.week)
        print("值班记录:", self.content)
        print("\n")


if __name__ == '__main__':
    print('PyCharm')
    # 读取当前目录下的excel文件（工作记录）
    excel_path = os.path.join(os.getcwd(), '2021年度值班工作记录表.xls')
    print("文件路径：", excel_path)
    excel_file = xlrd.open_workbook(excel_path)
    # 查看所有表格名字和个数
    names = excel_file.sheet_names()
    sheet_number = len(names)
    # 创建文件夹来存放结果文件
    if not os.path.exists('./月报'):  # 判断文件夹是否已经存在
        os.mkdir('./月报')
    if not os.path.exists('./周报'):  # 判断文件夹是否已经存在
        os.mkdir('./周报')
    # 打开周报模板
    '''excel_path_ZhouBao = os.path.join(os.getcwd(), 'pattern_week.xlsx')
    wb = openpyxl.load_workbook(excel_path_ZhouBao)
    sheet = wb['周报']
    sheet.title = '周报'
    sheet['C2'] = '终端运行室'''

    # 打开月报模板
    excel_path_YueBao = os.path.join(os.getcwd(), 'pattern_month.xlsx')
    wb2 = openpyxl.load_workbook(excel_path_YueBao)
    sheet = wb2['Sheet1']
    sheet.title = '月报'

    # Record类的数组，用来记录所有的值班记录
    records = []

    # 遍历所有表格，并将值班记录逐条加入record中
    for i in range(0, sheet_number):
        table = excel_file.sheets()[i]
        nrows = table.nrows
        rows = 6  # 值班记录结束的下标
        ncount = 6
        p = 0
        for k in range(6, nrows):
            stri = str(table.cell(k, 0).value)
            if stri.startswith('2'):
                rows = rows + 1
            else:
                break
        for j in range(6, rows):
            cur_date = str(int(table.cell(j, 0).value))  # 当前时间
            cur_week = str(table.cell(j, 1).value)  # 当前星期
            cur_content = str(table.cell(j, 4).value)  # 当前值班记录
            records.append(Record(cur_date, cur_week, cur_content))
    # for t in range(0,len(records)):
    #     print(records[t].week)
    nrecord = len(records)
    print("总共有" + str(nrecord) + "条记录")

    ### 值班记录统计完毕，开始做月报 ###
    ##声明各个值班模块内容以及时间
    str1 = []  # 甚高频、内话、记录仪
    str1_date = []
    str2 = []  # 自动化、ADS-B数据站
    str2_date = []
    str3 = []  # CDM、CNMS、数字空管
    str3_date = []
    str4 = []  # 晋江系统、A-CDM
    str4_date = []
    str5 = []  # 其他
    str5_date = []

    start_date = parse(records[0].date)  # 开始日期设置为records中的第一条记录
    content = ""  # 本月的内容汇总

    for i in range(0, nrecord):
        cur_date = parse(records[i].date)  # 当前遍历的位置的时间，当作指针使用
        cur_str = records[i].content  # 把当前记录的内容赋值给cur_str，方便后续使用
        if cur_date.day != 20:  # 统计20号之前的信息
            strlist = cur_str.split('\n')  # 内容按换行符分割，赋值黑strlist列表
            if len(strlist) < 5: # 没有内容则跳过继续执行扫描下条
                continue
            # 接下来处理数据，判断是否有内容记录，若有则记录到相应的str'i'中
            # 当前记录是否有甚高频、内话、记录仪内容 #
            if len(strlist[0]) > 14:
                str1.append(strlist[0][13:])
                str1_date.append(cur_date)
            # 当前记录是否有自动化、ADS-B数据站 #
            if len(strlist[1]) > 16:
                str2.append(strlist[1][15:])
                str2_date.append(cur_date)
            # 当前记录是否有CDM、CNMS、数字空管 #
            if len(strlist[2]) > 17:
                str3.append(strlist[2][16:])
                str3_date.append(cur_date)
            # 当前记录是否有晋江系统、A-CDM #
            if len(strlist[3]) > 14:
                str4.append(strlist[3][13:])
                str4_date.append(cur_date)
            # 当前记录是否有其他内容 #
            if len(strlist[4]) > 6:
                str5.append(strlist[4][5:])
                str5_date.append(cur_date)
        else:  # 统计到20号为止，输出本月内容到xlsx，重设开始时间再开始下一轮新的统计
            excel_time = str(start_date.year) + "年" + str(start_date.month) + "月" + str(start_date.day) + "日 ~ " + \
                         records[i - 1].date[0:4] + "年" + records[i - 1].date[4:6] + "月" + records[i - 1].date[6:] + "日"
            # print(excel_time)
            start_date = parse(records[i].date)
            content = "运行情况：\n" + "一、VHF、内话、记录仪\n"
            for j in range(0, len(str1)):
                content = content + str(str1_date[j].month) + "月" + str(str1_date[j].day) + "日  " + str1[j] + '\n'
            content = content + '\n' + "二、自动化、ADS-B数据站\n"
            for j in range(0, len(str2)):
                content = content + str(str2_date[j].month) + "月" + str(str2_date[j].day) + "日  " + str2[j] + '\n'
            content = content + '\n' + "三、CNMS、CDM、数字空管\n"
            for j in range(0, len(str3)):
                content = content + str(str3_date[j].month) + "月" + str(str3_date[j].day) + "日  " + str3[j] + '\n'
            content = content + '\n' + "四、晋江系统、A-CDM\n"
            for j in range(0, len(str4)):
                content = content + str(str4_date[j].month) + "月" + str(str4_date[j].day) + "日  " + str4[j] + '\n'
            content = content + '\n' + "五、其他\n"
            for j in range(0, len(str5)):
                content = content + str(str5_date[j].month) + "月" + str(str5_date[j].day) + "日  " + str5[j] + '\n'
            # print(content)
            sheet['B1'] = excel_time
            sheet['C4'] = content
            wb2.save('./月报/' + excel_time + '月报.xlsx')

            # 本月处理完毕，清空数据开始下个月统计
            str1.clear()
            str1_date.clear()
            str2.clear()
            str2_date.clear()
            str3.clear()
            str3_date.clear()
            str4.clear()
            str4_date.clear()
            str5.clear()
            str5_date.clear()
            content = ''
        # if cur_date.date

    ## 处理最后一轮情况：最后一条记录不是20号结尾
    '''
    有待完善
    '''
    print("月报已经成功生成！请认真核查月报内容，如有错误请手动修改哦")
    # print(str(start_date.day))
    # print(records[nrecord - 1].date[6:])
    # print(records[0].content)
    # strlist = records[0].content.split('\n')
    # print(strlist)
    # print(strlist[4][0:4])

    ## 月报完毕 ##

    ### 值班记录统计完毕，开始做周报 ###
    '''start_date = parse(records[0].date)  # 开始日期设置为records中的第一条记录
    content = ''  # 初始化内容板
    for i in range(0, nrecord):
        cell_date = parse(records[i].date)
        cell_time = str(cell_date.month) + '月' + str(cell_date.day) + '日' + '\n'  # 将当前位置的日期转化为xx月xx日格式
        content = content + cell_time + records[i].content
        if '星期四' in records[i].week:
            end_date = parse(records[i].date)
            sheet['C3'] = content
            sheet['J2'] = str(start_date.year) + '年' + str(start_date.month) + '月' + str(start_date.day) + '日至' + str(
                end_date.year) + '年' + str(end_date.month) + '月' + str(end_date.day) + '日'
            save_name = '周报初稿' + str(start_date.date()) + '到' + str(end_date.date()) + '.xlsx'
            wb.save(save_name)
            # 产生完一份周报，重新初始化内容板数据和起止时间
            content = ''
            start_date = parse(records[i + 1].date)
            end_date = ''
            continue

    # 处理最后一周的周报不是周四结尾的情况
    if '星期四' not in records[nrecord - 1].week:
        end_date = parse(records[nrecord - 1].date)
        for i in range(nrecord, nrecord - 7, -1):
            if '星期五' in records[nrecord - 1].week:
                start_date = parse(records[i - 1].date)
                break
        sheet['C3'] = content
        sheet['J2'] = str(start_date.year) + '年' + str(start_date.month) + '月' + str(start_date.day) + '日至' + str(
            end_date.year) + '年' + str(end_date.month) + '月' + str(end_date.day) + '日'
        save_name = '周报初稿' + str(start_date.date()) + '到' + str(end_date.date()) + '.xlsx'
        wb.save(save_name)
        '''
