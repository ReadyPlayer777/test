# 这是一个示例 Python 脚本。
'''
单元格格式说明
C2:科室名称 J2:运行日期 C3:本周运行和管理 C46:定期任务和执行 C51:交办任务 C55:下周计划 C59:点评及任务布置 C63:文件传达及学习
'''
import xlrd
import xlwt
import openpyxl
import os
from dateutil.parser import parse


# 值班记录类
class Record():
    def __init__(self, date, week, content):
        self.date = date
        self.week = week
        self.content = content

    def printinfo(self):
        print("日期:", self.date)
        print("星期:", self.week)
        print("值班记录:", self.content)
        print("\n")


if __name__ == '__main__':
    print('PyCharm')
    # 读取当前目录下的excel文件（工作记录）
    excel_path = os.path.join(os.getcwd(), '2021年度值班工作记录表.xls')
    print("文件路径：", excel_path)
    excel_file = xlrd.open_workbook(excel_path)
    # 查看所有表格名字和个数
    names = excel_file.sheet_names()
    sheet_number = len(names)

    if not os.path.exists('./周报'):  # 判断文件夹是否已经存在
        os.mkdir('./周报')

    # 打开周报模板
    excel_path_ZhouBao = os.path.join(os.getcwd(), 'pattern_week.xlsx')
    wb = openpyxl.load_workbook(excel_path_ZhouBao)
    sheet = wb['周报']
    sheet.title = '周报'
    sheet['C2'] = '终端运行室'

    # Record类的数组，用来记录所有的值班记录
    records = []

    # 遍历所有表格，并将值班记录逐条加入record中
    for i in range(0, sheet_number):
        table = excel_file.sheets()[i]
        nrows = table.nrows
        rows = 6  # 值班记录结束的下标
        ncount = 6
        p = 0
        for k in range(6, nrows):
            stri = str(table.cell(k, 0).value)
            if stri.startswith('2'):
                rows = rows + 1
            else:
                break
        for j in range(6, rows):
            cur_date = str(int(table.cell(j, 0).value))  # 当前时间
            cur_week = str(table.cell(j, 1).value)  # 当前星期
            cur_content = str(table.cell(j, 4).value)  # 当前值班记录
            records.append(Record(cur_date, cur_week, cur_content))
    # for t in range(0,len(records)):
    #     print(records[t].week)
    nrecord = len(records)
    print("总共有" + str(nrecord) + "条记录")

    ### 值班记录统计完毕，开始做周报 ###
    start_date = parse(records[0].date)  # 开始日期设置为records中的第一条记录
    content = ''  # 初始化内容板
    for i in range(0, nrecord):
        cell_date = parse(records[i].date)
        cell_time = str(cell_date.month) + '月' + str(cell_date.day) + '日' + '\n'  # 将当前位置的日期转化为xx月xx日格式
        content = content + cell_time + records[i].content
        if '星期四' in records[i].week:
            end_date = parse(records[i].date)
            sheet['C3'] = content
            sheet['J2'] = str(start_date.year) + '年' + str(start_date.month) + '月' + str(start_date.day) + '日至' + str(
                end_date.year) + '年' + str(end_date.month) + '月' + str(end_date.day) + '日'
            save_name = '周报初稿' + str(start_date.date()) + '到' + str(end_date.date()) + '.xlsx'
            wb.save('./周报/' + save_name)
            # 产生完一份周报，重新初始化内容板数据和起止时间
            content = ''
            start_date = parse(records[i + 1].date)
            end_date = ''
            continue

    # 处理最后一周的周报不是周四结尾的情况
    if '星期四' not in records[nrecord - 1].week:
        end_date = parse(records[nrecord - 1].date)
        for i in range(nrecord, nrecord - 7, -1):
            if '星期五' in records[nrecord - 1].week:
                start_date = parse(records[i - 1].date)
                break
        sheet['C3'] = content
        sheet['J2'] = str(start_date.year) + '年' + str(start_date.month) + '月' + str(start_date.day) + '日至' + str(
            end_date.year) + '年' + str(end_date.month) + '月' + str(end_date.day) + '日'
        save_name = '周报初稿' + str(start_date.date()) + '到' + str(end_date.date()) + '.xlsx'
        wb.save('./周报/' + save_name)

    print("周报已经成功生成！请认真核查月报内容，如有错误请手动修改哦！！！")
